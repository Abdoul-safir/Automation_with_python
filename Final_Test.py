"""The openpyxl library has been downloaded for its fonctionalities to be usable in this code"""
"""Importing foction workbook from openpyxl library"""
from openpyxl.workbook import Workbook
"""Importing foction load_workbook from openpyxl library that permits to access the excel sheet"""
from openpyxl import load_workbook
"""Importing csv library which permits to carry out operations on csv files"""
import csv

"""using the load_workbook to access the excel sheet which is assigned to the variable wb"""
wb = load_workbook('employeedata.xlsx')
"""Using the active fonction to use the sheet containing our data"""
ws = wb.active

"""for loop that permits to move accross the columns of the row to be edited"""
for row in ws.rows:
        row[3].value
        Abd = row[3].value
        """The replace fonction searches the string to be replaced and replaces it """
        row[3].value = Abd.replace("@helpinghands.cm","handsinhands.org")
"""saves the final version into Output.xlsx"""   
wb.save('Output.xlsx')
    
  
outfile = open('CSVOutput.csv ', 'w')
with open('employeedata.csv', 'r') as csv_file:
      csv_reader = csv.reader(csv_file)
      header = next(csv_reader)
      for row in csv_reader:
        NAME = row[0]
        L_NAME = row[1]
        CONTACT =row[2]
        EMAIL = row[3]
        ADDRESS =row[4]
        if 'helpinghands.cm' in EMAIL:
            new_email=(EMAIL).replace('helpinghands.cm','handsinhands.org')
        line = "{},{},{},{}\n".format(NAME,L_NAME, CONTACT,new_email, ADDRESS)
        outfile.write(line)
    
outfile.close()
        
